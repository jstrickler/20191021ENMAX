#!/usr/bin/env python
import openpyxl as px


def main():
    """program entry point"""
    wb = px.load_workbook('../DATA/presidents.xlsx')
    ws = wb['US Presidents']
    ws2 = wb.create_sheet('President Names')

    save_range_to_new_sheet(ws, ws2)

    wb.save()


def save_range_to_new_sheet(ws, ws2):
    """Print first and last names of all presidents"""
    pres_range = ws['B2':'C46']  # cell range

    for i, row in enumerate(pres_range, 1):
        for j, col in enumerate(row, 1):
            ws2.cell(row=i, column=j).value = ws.cell(row=i, column=j)

if __name__ == '__main__':
    main()
